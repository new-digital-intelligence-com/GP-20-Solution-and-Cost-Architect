"""
Commercial pricing form — offering-agnostic.

This writer has no idea whether it is pricing a network, a device fleet or a
service desk. It consumes only the contract shape, which is what lets one
writer serve every pack.

Usage:  python write_pricing_form.py result.json [output.xlsx]
"""

from __future__ import annotations

# --- Windows console safety -------------------------------------------------
import sys as _sys
for _s in (_sys.stdout, _sys.stderr):
    try:
        if _s and getattr(_s, "encoding", "").lower().replace("-", "") != "utf8":
            _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
# ---------------------------------------------------------------------------

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NAVY, ACCENT = "1F3556", "2E6E8E"
TITLE = Font(bold=True, size=14, color="FFFFFF")
HDR = Font(bold=True, size=10, color="FFFFFF")
SECT = Font(bold=True, size=11, color=NAVY)
BOLD = Font(bold=True, size=10)
BASE = Font(size=10)
BIG = Font(bold=True, size=13, color=NAVY)
MUTED = Font(size=9, italic=True, color="6B7785")
WARN = Font(size=9, bold=True, color="B4472B")

F_NAVY = PatternFill("solid", fgColor=NAVY)
F_ACCENT = PatternFill("solid", fgColor=ACCENT)
F_SECT = PatternFill("solid", fgColor="E8EDF3")
F_TOTAL = PatternFill("solid", fgColor="EAF3EC")
F_WARN = PatternFill("solid", fgColor="FDEBE0")

THIN = Side(style="thin", color="C9D2DC")
BOX = Border(THIN, THIN, THIN, THIN)
NUM2 = "#,##0.00"

FMT = {"int": "#,##0", "float1": "#,##0.0", "float2": "#,##0.00", "text": None}


def build(result: dict, out_path: str) -> str:
    meta = result["meta"]
    money = f'"{meta["symbol"]}"#,##0'
    money2 = f'"{meta["symbol"]}"#,##0.00'

    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing Form"
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCD", [48, 19, 19, 40]):
        ws.column_dimensions[col].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = ws.page_margins.right = 0.4

    dep, run, summ = result["deployment"], result["run"], result["summary"]
    r = 1

    def title(text, sub=""):
        nonlocal r
        c = ws.cell(r, 1, text)
        c.font, c.fill = TITLE, F_NAVY
        for i in range(2, 5):
            ws.cell(r, i).fill = F_NAVY
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        if sub:
            ws.cell(r, 1, sub).font = MUTED
            r += 1

    def section(text):
        nonlocal r
        r += 1
        for i in range(1, 5):
            ws.cell(r, i).fill = F_SECT
        ws.cell(r, 1, text).font = SECT
        r += 1

    def kv(label, value, fmt=None, note=""):
        nonlocal r
        ws.cell(r, 1, label).font = BASE
        c = ws.cell(r, 2, value)
        c.font = BOLD
        if fmt:
            c.number_format = fmt
        if note:
            ws.cell(r, 4, note).font = MUTED
        r += 1

    def head(headers):
        nonlocal r
        for i, h in enumerate(headers, 1):
            c = ws.cell(r, i, h)
            c.font, c.fill = HDR, F_ACCENT
            c.alignment = Alignment(horizontal="center" if i > 1 else "left",
                                    wrap_text=True)
        r += 1

    def row(cells, fmts=None, bold=False, fill=None):
        nonlocal r
        for i, v in enumerate(cells, 1):
            c = ws.cell(r, i, v)
            c.font = BOLD if (bold or i == 1) else BASE
            c.border = BOX
            if fill:
                c.fill = fill
            if fmts and i - 1 < len(fmts) and fmts[i - 1]:
                c.number_format = fmts[i - 1]
                c.alignment = Alignment(horizontal="right")
        r += 1

    def total(label, value, fmt=None):
        nonlocal r
        ws.cell(r, 1, label).font = BIG
        c = ws.cell(r, 2, value)
        c.font = BIG
        c.number_format = fmt or money
        for i in (1, 2):
            ws.cell(r, i).fill = F_TOTAL
            ws.cell(r, i).border = BOX
        r += 1

    # ---------------------------------------------------------------
    title("COMMERCIAL PRICING FORM", meta["offering_name"])
    r += 1
    kv("Client", meta["client_name"])
    kv("RFP reference", meta["rfp_ref"])
    kv("Offering", meta["offering_name"])
    kv("Currency", meta["currency"])
    kv("Contract term", f"{summ['term_years']} years")

    ws.cell(r, 1, meta["disclaimer"]).font = WARN
    for i in range(1, 5):
        ws.cell(r, i).fill = F_WARN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1

    # A. Scope
    section("A.  SOLUTION SUMMARY")
    for item in result["scope"]["headline"]:
        kv(item["label"], item["value"], FMT.get(item["format"], "#,##0"))
    for item in result["service"]:
        kv(item["label"], item["value"])

    # B. One-off
    section("B.  ONE-OFF CHARGES")
    head(["Element", "Effort (days)", "Amount", "Notes"])
    for role, days in dep["days_by_role"].items():
        if role == "total":
            continue
        row([role.replace("_", " ").title(), days, "", ""], [None, NUM2, money])
    row(["Total effort", dep["days_by_role"].get("total", 0), "", ""],
        [None, NUM2, money], bold=True)
    for line in dep["cost_lines"]:
        row([line["label"], "", line["amount"], ""], [None, None, money])
    total("TOTAL ONE-OFF CHARGE", dep["price"])

    # C. Resources
    if run["resources"]:
        section("C.  RESOURCE MODEL")
        head(["Location", "Role", "FTE", "Sizing driver"])
        for res in run["resources"]:
            row([res["location"], res["role"], res["fte"], res.get("driver", "")],
                [None, None, NUM2])
        row(["Total", "", run["total_fte"], ""], [None, None, NUM2], bold=True)

    if run.get("insight"):
        ws.cell(r, 1, "Architect's note").font = BOLD
        r += 1
        c = ws.cell(r, 1, run["insight"])
        c.font = Font(size=9, italic=True, color=NAVY)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)
        ws.row_dimensions[r].height = 26
        r += 3

    # D. Recurring
    section("D.  RECURRING ANNUAL CHARGES")
    head(["Element", "Amount", "Notes"])
    for line in run["cost_lines"]:
        note = "Credit" if line["amount"] < 0 else ""
        row([line["label"], line["amount"], note], [None, money])
    total("ANNUAL CHARGE (YEAR 1)", run["price_pa"])
    total("MONTHLY CHARGE (YEAR 1)", run["price_pm"])
    for u in summ.get("unit_metrics", []):
        total(u["label"].upper(), u["value"], money2)

    # E. Schedule
    section("E.  PRICE SCHEDULE BY CONTRACT YEAR")
    head(["Contract year", "Annual charge", "Notes"])
    for y in summ["yearly"]:
        row([f"Year {y['year']}", y["price"],
             "Base year" if y["year"] == 1 else "Indexed"], [None, money])

    # F. TCV — shown as a cost-to-price bridge, because the first question
    # finance asks is what the margin is being taken on.
    section("F.  TOTAL CONTRACT VALUE")
    kv("One-off charges", dep["price"], money)
    kv("Recurring over term", round(summ["tcv"] - dep["price"], 2), money)
    r += 1
    kv("Delivery cost over term", summ["total_cost"], money,
       "Before pricing adjustment")
    kv("Margin", summ["margin_value"], money,
       f"{summ['margin_pct_effective']:.1%} of contract value")
    total(f"TOTAL CONTRACT VALUE ({summ['term_years']} YEARS)", summ["tcv"])

    # G. Assumptions
    section("G.  ASSUMPTIONS AND CLARIFICATIONS")
    ws.cell(r, 1, "Parameters resting on model defaults are listed first — these "
                  "require confirmation before the price is issued.").font = MUTED
    r += 1
    head(["Parameter", "Value", "Source", "Note"])
    flags = {f["parameter"] for f in result.get("review_flags", [])}
    entries = result.get("assumptions", [])
    ordered = ([a for a in entries if a["parameter"] in flags]
               + [a for a in entries if a["parameter"] not in flags])
    for a in ordered:
        val = a["value"]
        if isinstance(val, dict):
            val = ", ".join(f"{k}={v}" for k, v in val.items())
        row([a["parameter"], str(val)[:40], a["source"], a.get("note", "")],
            fill=F_WARN if a["parameter"] in flags else None)

    r += 1
    ws.cell(r, 1, meta["disclaimer"]).font = WARN

    _bom_sheet(wb, result)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _bom_sheet(wb: Workbook, result: dict) -> None:
    """Bill of materials — quantities at cost, never price.

    Ships in the same workbook as the bill of services because that is how the
    two are read: what is being bought, then what it costs to run. Every line
    states a quantity, a unit and a unit cost, so a reviewer can challenge any
    number without opening the model.
    """
    bom = result.get("bom") or []
    if not bom:
        return

    meta = result["meta"]
    money = f'"{meta["symbol"]}"#,##0'
    money4 = f'"{meta["symbol"]}"#,##0.00'

    ws = wb.create_sheet("Bill of Materials")
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEF", [42, 12, 22, 16, 18, 46]):
        ws.column_dimensions[col].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    r = 1
    c = ws.cell(r, 1, "BILL OF MATERIALS")
    c.font, c.fill = TITLE, F_NAVY
    for i in range(2, 7):
        ws.cell(r, i).fill = F_NAVY
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 1, f"{meta['offering_name']} — {meta['client_name']}").font = MUTED
    r += 1
    ws.cell(r, 1, "Costs shown are before margin, contingency and any other "
                  "pricing adjustment.").font = MUTED
    r += 1

    ws.cell(r, 1, meta["disclaimer"]).font = WARN
    for i in range(1, 7):
        ws.cell(r, i).fill = F_WARN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2

    grand = 0.0
    for phase, heading in (("one-off", "ONE-OFF — PURCHASED AT DEPLOYMENT"),
                           ("recurring", "RECURRING — ANNUAL, YEAR ONE")):
        rows = [b for b in bom if b["phase"] == phase]
        if not rows:
            continue

        for i in range(1, 7):
            ws.cell(r, i).fill = F_SECT
        ws.cell(r, 1, heading).font = SECT
        r += 1

        for i, h in enumerate(["Item", "Quantity", "Unit", "Unit cost",
                               "Extended cost", "Basis"], 1):
            cell = ws.cell(r, i, h)
            cell.font, cell.fill = HDR, F_ACCENT
            cell.alignment = Alignment(horizontal="center" if 1 < i < 6 else "left",
                                       wrap_text=True)
        r += 1

        subtotal = 0.0
        for cat in dict.fromkeys(b.get("category", "") for b in rows):
            in_cat = [b for b in rows if b.get("category", "") == cat]
            if cat and len(dict.fromkeys(b.get("category", "") for b in rows)) > 1:
                ws.cell(r, 1, cat).font = Font(bold=True, size=9, color=ACCENT)
                r += 1
            for b in in_cat:
                vals = [b["item"], round(b["qty"], 1), b["unit"],
                        b["unit_cost"], b["extended_cost"], b.get("note", "")]
                for i, v in enumerate(vals, 1):
                    cell = ws.cell(r, i, v)
                    cell.font = BASE
                    cell.border = BOX
                    if i == 2:
                        cell.number_format = "#,##0.0"
                        cell.alignment = Alignment(horizontal="right")
                    elif i == 4:
                        cell.number_format = money4
                        cell.alignment = Alignment(horizontal="right")
                    elif i == 5:
                        cell.number_format = money
                        cell.alignment = Alignment(horizontal="right")
                    elif i == 6:
                        cell.font = MUTED
                subtotal += b["extended_cost"]
                r += 1

        ws.cell(r, 1, f"Subtotal — {phase}").font = BIG
        cell = ws.cell(r, 5, round(subtotal, 2))
        cell.font, cell.number_format = BIG, money
        for i in (1, 5):
            ws.cell(r, i).fill = F_TOTAL
            ws.cell(r, i).border = BOX
        r += 2
        grand += subtotal

    ws.cell(r, 1, "Materials are a component of the cost lines in the pricing "
                  "form, not a separate charge — labour, travel and overhead sit "
                  "alongside them.").font = MUTED
    r += 1
    ws.cell(r, 1, meta["disclaimer"]).font = WARN


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    src = sys.argv[1]
    result = json.load(sys.stdin if src == "-" else open(src, encoding="utf-8"))
    out = sys.argv[2] if len(sys.argv) > 2 else "Pricing_Form.xlsx"
    print(f"Written: {build(result, out)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
