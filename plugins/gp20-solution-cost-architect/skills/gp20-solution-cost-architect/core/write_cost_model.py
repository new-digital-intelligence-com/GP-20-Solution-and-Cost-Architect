"""
Per-deal cost model workbook — offering-agnostic.

This is the audit artefact. The pricing form says what the client is charged;
this says how that number was arrived at, from the inputs down, with the
margin visible as its own step rather than baked into every line.

It is a *live* workbook, not a dump. Costs are values, because the pack owns
the arithmetic and nothing here may quietly re-derive them. Everything
downstream of cost — margin, price, indexation, the term schedule, contract
value — is an Excel formula reading from named input cells. Finance can flex
the margin or the indexation and watch the schedule move, without a Python
round trip and without anyone editing a cost.

That split is the whole design: the pack owns cost, the workbook owns pricing.

Usage:  python write_cost_model.py result.json [output.xlsx]
"""

from __future__ import annotations

# --- Windows console safety -------------------------------------------------
import sys as _sys
for _s in (_sys.stdout, _sys.stderr):
    try:
        if _s and getattr(_s, "encoding", "").lower().replace("-", "") != "utf8":
            _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
# ---------------------------------------------------------------------------

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY, ACCENT = "1F3556", "2E6E8E"
TITLE = Font(bold=True, size=14, color="FFFFFF")
HDR = Font(bold=True, size=10, color="FFFFFF")
SECT = Font(bold=True, size=11, color=NAVY)
BOLD = Font(bold=True, size=10)
BASE = Font(size=10)
BIG = Font(bold=True, size=12, color=NAVY)
MUTED = Font(size=9, italic=True, color="6B7785")
WARN = Font(size=9, bold=True, color="B4472B")
INPUT_FONT = Font(bold=True, size=10, color="7A4A0C")

F_NAVY = PatternFill("solid", fgColor=NAVY)
F_ACCENT = PatternFill("solid", fgColor=ACCENT)
F_SECT = PatternFill("solid", fgColor="E8EDF3")
F_TOTAL = PatternFill("solid", fgColor="EAF3EC")
F_WARN = PatternFill("solid", fgColor="FDEBE0")
F_INPUT = PatternFill("solid", fgColor="FDF3E4")

THIN = Side(style="thin", color="C9D2DC")
BOX = Border(THIN, THIN, THIN, THIN)
PCT = "0.00%"
FMT = {"int": "#,##0", "float1": "#,##0.0", "float2": "#,##0.00",
       "money": "#,##0", "text": None}

SOURCE_LABEL = {"rfp": "RFP", "user": "User", "derived": "Derived",
                "default": "Model default"}


def build(result: dict, out_path: str) -> str:
    meta, summ = result["meta"], result["summary"]
    dep, run = result["deployment"], result["run"]
    money = f'"{meta["symbol"]}"#,##0'
    money2 = f'"{meta["symbol"]}"#,##0.00'

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Model"
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEF", [46, 18, 18, 18, 18, 34]):
        ws.column_dimensions[col].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    r = 1

    def band(text, sub=""):
        nonlocal r
        ws.cell(r, 1, text).font = TITLE
        for i in range(1, 7):
            ws.cell(r, i).fill = F_NAVY
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        if sub:
            ws.cell(r, 1, sub).font = MUTED
            r += 1

    def section(text, note=""):
        nonlocal r
        r += 1
        for i in range(1, 7):
            ws.cell(r, i).fill = F_SECT
        ws.cell(r, 1, text).font = SECT
        r += 1
        if note:
            ws.cell(r, 1, note).font = MUTED
            r += 1

    def head(headers):
        nonlocal r
        for i, h in enumerate(headers, 1):
            c = ws.cell(r, i, h)
            c.font, c.fill = HDR, F_ACCENT
            c.alignment = Alignment(horizontal="center" if i > 1 else "left",
                                    wrap_text=True)
        r += 1

    def line(label, value, fmt=None, note="", col=2):
        """A plain value row. Returns the cell reference for later formulas."""
        nonlocal r
        ws.cell(r, 1, label).font = BASE
        c = ws.cell(r, col, value)
        c.font, c.border = BASE, BOX
        if fmt:
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right")
        if note:
            ws.cell(r, 6, note).font = MUTED
        ref = f"{get_column_letter(col)}{r}"
        r += 1
        return ref

    def input_cell(label, value, fmt, note=""):
        """An amber cell finance is meant to change."""
        nonlocal r
        ws.cell(r, 1, label).font = BOLD
        c = ws.cell(r, 2, value)
        c.font, c.fill, c.border, c.number_format = INPUT_FONT, F_INPUT, BOX, fmt
        c.alignment = Alignment(horizontal="right")
        if note:
            ws.cell(r, 6, note).font = MUTED
        ref = f"B{r}"
        r += 1
        return ref

    def total_row(label, formula, fmt=None, note=""):
        nonlocal r
        ws.cell(r, 1, label).font = BIG
        c = ws.cell(r, 2, formula)
        c.font, c.number_format, c.border = BIG, fmt or money, BOX
        c.alignment = Alignment(horizontal="right")
        for i in (1, 2):
            ws.cell(r, i).fill = F_TOTAL
        if note:
            ws.cell(r, 6, note).font = MUTED
        ref = f"B{r}"
        r += 1
        return ref

    # -----------------------------------------------------------------
    band("COST MODEL", f"{meta['offering_name']} — {meta['client_name']} · "
                       f"{meta['rfp_ref']}")
    ws.cell(r, 1, meta["disclaimer"]).font = WARN
    for i in range(1, 7):
        ws.cell(r, i).fill = F_WARN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    # A — the levers
    section("A.  COMMERCIAL PARAMETERS",
            "Amber cells are inputs. Everything below cost is a formula reading "
            "from them — change a lever and the schedule follows.")
    margin_ref = input_cell("Target margin", summ["margin_pct_effective"], PCT,
                            "Applied to cost to reach price")
    index_ref = input_cell("Annual indexation", summ.get("indexation_pct", 0.0),
                           PCT, "Applied from year two")
    term_ref = input_cell("Contract term (years)", summ["term_years"], "#,##0")

    # B — what is being priced
    section("B.  SOLUTION SCOPE")
    for item in result["scope"]["headline"]:
        line(item["label"], item["value"], FMT.get(item["format"], "#,##0"))
    for item in result["service"]:
        line(item["label"], item["value"])

    # C — how the size was arrived at
    drivers = {k: v for k, v in (run.get("drivers") or {}).items()
               if isinstance(v, (int, float))}
    days = {k: v for k, v in dep.get("days_by_role", {}).items() if k != "total"}
    if drivers or days:
        section("C.  SIZING DRIVERS",
                "The intermediate quantities the cost is built from — the first "
                "place to look when a total looks wrong.")
        for k, v in days.items():
            line(f"Deployment effort — {k.replace('_', ' ')}", round(v, 1),
                 "#,##0.0", "days")
        if days:
            line("Deployment effort — total",
                 round(dep["days_by_role"].get("total", 0), 1), "#,##0.0", "days")
        for k, v in drivers.items():
            line(k.replace("_", " ").capitalize(), round(float(v), 2), "#,##0.00")

    # D — who delivers it
    if run.get("resources"):
        section("D.  RESOURCE MODEL")
        head(["Location", "Role", "FTE", "", "", "Sizing driver"])
        for res in run["resources"]:
            ws.cell(r, 1, res["location"]).font = BASE
            ws.cell(r, 2, res["role"]).font = BASE
            c = ws.cell(r, 3, round(res["fte"], 2))
            c.font, c.number_format, c.border = BASE, "#,##0.00", BOX
            ws.cell(r, 6, res.get("driver", "")).font = MUTED
            r += 1
        ws.cell(r, 1, "Total FTE").font = BOLD
        c = ws.cell(r, 3, run["total_fte"])
        c.font, c.number_format = BOLD, "#,##0.00"
        r += 1

    # E — one-off
    section("E.  ONE-OFF COST BUILD",
            "Cost lines are produced by the model and are not editable here.")
    head(["Cost element", "Amount", "", "", "", "Notes"])
    first = r
    for cl in dep["cost_lines"]:
        line(cl["label"], cl["amount"], money)
    dep_cost = total_row("Total one-off cost",
                         f"=SUM(B{first}:B{r - 1})", money)
    dep_price = total_row("One-off price",
                          f"={dep_cost}/(1-{margin_ref})", money,
                          "Cost divided by one minus margin")

    # F — recurring
    section("F.  RECURRING COST BUILD — YEAR ONE",
            "Year one, before indexation and before any consumption ramp.")
    head(["Cost element", "Amount", "", "", "", "Notes"])
    first = r
    for cl in run["cost_lines"]:
        line(cl["label"], cl["amount"], money,
             "Credit" if cl["amount"] < 0 else "")
    run_cost = total_row("Total annual cost (year 1)",
                         f"=SUM(B{first}:B{r - 1})", money)
    run_price = total_row("Annual price (year 1)",
                          f"={run_cost}/(1-{margin_ref})", money)
    total_row("Monthly price (year 1)", f"={run_price}/12", money2)

    # G — the schedule
    profile = summ.get("year_profile") or [1.0] * summ["term_years"]
    section("G.  TERM SCHEDULE",
            "Consumption profile multiplied by indexation. Year one is the base "
            "for both.")
    head(["Contract year", "Consumption factor", "Annual cost", "Annual price",
          "", "Basis"])
    sched_first = r
    for y in range(1, summ["term_years"] + 1):
        factor = profile[y - 1] if y - 1 < len(profile) else profile[-1]
        ws.cell(r, 1, f"Year {y}").font = BASE
        c = ws.cell(r, 2, round(factor, 6))
        c.font, c.number_format, c.border = BASE, "#,##0.0000", BOX
        for col, base_ref in ((3, run_cost), (4, run_price)):
            c = ws.cell(r, col,
                        f"={base_ref}*B{r}*(1+{index_ref})^({y}-1)")
            c.font, c.number_format, c.border = BASE, money, BOX
            c.alignment = Alignment(horizontal="right")
        ws.cell(r, 6, "Base year" if y == 1 else f"Indexed {y - 1}x").font = MUTED
        r += 1
    sched_last = r - 1

    # H — the bridge
    section("H.  CONTRACT VALUE",
            "The only place margin appears as a number rather than a divisor.")
    tcc = total_row("Total delivery cost over term",
                    f"={dep_cost}+SUM(C{sched_first}:C{sched_last})", money,
                    "What the business spends")
    tcv = total_row("Total contract value",
                    f"={dep_price}+SUM(D{sched_first}:D{sched_last})", money,
                    "What the client is charged")
    total_row("Margin", f"={tcv}-{tcc}", money)
    total_row("Margin as % of contract value", f"=({tcv}-{tcc})/{tcv}", PCT)

    # I — provenance
    section("I.  INPUT REGISTER",
            "Every parameter and where it came from. Model defaults are listed "
            "first — those are the ones still to be confirmed.")
    head(["Parameter", "Value", "Source", "", "", "Note"])
    flags = {f["parameter"] for f in result.get("review_flags", [])}
    entries = result.get("assumptions", [])
    ordered = ([a for a in entries if a["parameter"] in flags]
               + [a for a in entries if a["parameter"] not in flags])
    for a in ordered:
        val = a["value"]
        if isinstance(val, dict):
            val = ", ".join(f"{k}={v}" for k, v in val.items())
        fill = F_WARN if a["parameter"] in flags else None
        for i, v in enumerate([a["parameter"], str(val)[:46],
                               SOURCE_LABEL.get(a["source"], a["source"])], 1):
            c = ws.cell(r, i, v)
            c.font, c.border = BASE, BOX
            if fill:
                c.fill = fill
        ws.cell(r, 6, a.get("note", "")).font = MUTED
        r += 1

    if run.get("insight"):
        r += 1
        ws.cell(r, 1, "Architect's note").font = BOLD
        r += 1
        c = ws.cell(r, 1, run["insight"])
        c.font = Font(size=9, italic=True, color=NAVY)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
        r += 3

    r += 1
    ws.cell(r, 1, meta["disclaimer"]).font = WARN

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    src = sys.argv[1]
    result = json.load(sys.stdin if src == "-" else open(src, encoding="utf-8"))
    out = sys.argv[2] if len(sys.argv) > 2 else "Cost_Model.xlsx"
    print(f"Written: {build(result, out)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
