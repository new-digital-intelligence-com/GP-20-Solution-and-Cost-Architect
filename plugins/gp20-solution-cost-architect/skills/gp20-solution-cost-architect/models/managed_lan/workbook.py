"""
Generate WLAN_Cost_Estimator.xlsx — a standalone, formula-driven Excel model.

Pack extra, not part of the contract. A workbook only makes sense for an
offering a human wants to poke at directly; not every pack will have one. It
lives here rather than in core precisely because it is offering-specific.

The Rates sheet is written from rates.py, so the workbook and the Python
engine share one source of truth for constants and cannot drift.
Every calculation on Model/Output is a live Excel formula: a Solution Shaper
can open the file, change an input, and watch it recalculate without Claude.

Usage:  python3 build_workbook.py [output_path]
"""



from __future__ import annotations

# --- Windows console safety -------------------------------------------------
# Windows defaults stdout to cp1252, which cannot encode the currency symbols,
# em-dashes and comparison operators this model emits. Force UTF-8 so the
# caller never has to remember `python -X utf8`.
import sys as _sys
for _s in (_sys.stdout, _sys.stderr):
    try:
        if _s and getattr(_s, "encoding", "").lower().replace("-", "") != "utf8":
            _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
# ---------------------------------------------------------------------------


import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

import sys as _s2
from pathlib import Path as _P
_s2.path.insert(0, str(_P(__file__).resolve().parent.parent.parent))
from models.managed_lan import rates as R

# --- Styling ---------------------------------------------------------------
NAVY = "1F3556"
ACCENT = "2E6E8E"
INPUT_FILL = PatternFill("solid", fgColor="FFF6DA")
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor="E8EDF3")
OUT_FILL = PatternFill("solid", fgColor="EAF3EC")
WARN_FILL = PatternFill("solid", fgColor="FDEBE0")

TITLE_F = Font(bold=True, size=14, color="FFFFFF")
HDR_F = Font(bold=True, size=10, color="FFFFFF")
SUB_F = Font(bold=True, size=10, color=NAVY)
BOLD = Font(bold=True, size=10)
BASE = Font(size=10)
MUTED = Font(size=9, italic=True, color="6B7785")

THIN = Side(style="thin", color="C9D2DC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GBP = '"£"#,##0'
GBP2 = '"£"#,##0.00'
NUM1 = "#,##0.0"
NUM2 = "#,##0.00"
PCT = "0.0%"

names: dict[str, str] = {}   # defined-name -> absolute address


def _abs(sheet: str, row: int, col: int = 2) -> str:
    return f"{sheet}!${get_column_letter(col)}${row}"


def _first_row(range_ref: str) -> int:
    """'Input!$B$20:$B$24' -> 20"""
    import re
    m = re.search(r"\$[A-Z]+\$(\d+)", range_ref)
    if not m:
        raise ValueError(f"Cannot parse first row from {range_ref!r}")
    return int(m.group(1))


# ===========================================================================
# Rates sheet
# ===========================================================================

def build_rates(wb: Workbook):
    ws = wb.create_sheet("Rates")
    ws.column_dimensions["A"].width = 34
    for c in "BCDEF":
        ws.column_dimensions[c].width = 15

    ws["A1"] = "RATE CARD — ILLUSTRATIVE / SYNTHETIC DATA"
    ws["A1"].font = TITLE_F
    ws["A1"].fill = HDR_FILL
    ws.merge_cells("A1:F1")
    ws["A2"] = "Not NSC pricing. Replace this sheet with the real rate card; no formula changes required."
    ws["A2"].font = MUTED

    r = 4

    def section(title, headers):
        nonlocal r
        ws.cell(r, 1, title).font = SUB_F
        ws.cell(r, 1).fill = SUB_FILL
        for i in range(2, len(headers) + 1):
            ws.cell(r, i).fill = SUB_FILL
        r += 1
        for i, h in enumerate(headers, start=1):
            c = ws.cell(r, i, h)
            c.font = HDR_F
            c.fill = PatternFill("solid", fgColor=ACCENT)
            c.alignment = Alignment(horizontal="center")
        r += 1
        return r

    # --- Country rates ---
    section("COUNTRY RATES (loaded day rates, GBP)",
            ["Country", "Field", "NOC", "Service Desk", "SDM", "Travel/incident"])
    country_start = r
    for c in R.SUPPORTED_COUNTRIES:
        v = R.COUNTRY_RATES[c]
        ws.cell(r, 1, c).font = BOLD
        for i, k in enumerate(["field", "noc", "service_desk", "sdm", "travel"], start=2):
            cell = ws.cell(r, i, v[k])
            cell.number_format = GBP
            cell.font = BASE
        r += 1
    country_end = r - 1
    names["CountryRates"] = f"Rates!$A${country_start}:$F${country_end}"
    names["FieldRates"] = f"Rates!$B${country_start}:$B${country_end}"
    names["TravelRates"] = f"Rates!$F${country_start}:$F${country_end}"
    names["SdmRates"] = f"Rates!$E${country_start}:$E${country_end}"
    r += 1

    # --- Site bands ---
    section("SITE BANDS", ["Band", "APs per site", "Users per site"])
    band_start = r
    for b in R.SITE_BANDS:
        ws.cell(r, 1, b).font = BOLD
        ws.cell(r, 2, R.APS_PER_SITE_BAND[b]).font = BASE
        ws.cell(r, 3, R.USERS_PER_SITE_BAND[b]).font = BASE
        r += 1
    band_end = r - 1
    names["ApsPerSite"] = f"Rates!$B${band_start}:$B${band_end}"
    names["UsersPerSite"] = f"Rates!$C${band_start}:$C${band_end}"
    r += 1

    # --- SLA tiers ---
    section("SLA TIERS", ["Tier", "FTE multiplier", "On-site response", "Triggers floor"])
    sla_start = r
    for t, m in R.SLA_MULTIPLIER.items():
        ws.cell(r, 1, t).font = BOLD
        ws.cell(r, 2, m).font = BASE
        ws.cell(r, 3, R.SLA_RESPONSE[t]).font = BASE
        ws.cell(r, 4, 1 if t in R.ON_SITE_RESPONSE_TIERS else 0).font = BASE
        r += 1
    names["SlaTable"] = f"Rates!$A${sla_start}:$D${r-1}"
    r += 1

    # --- Coverage ---
    section("COVERAGE WINDOWS", ["Coverage", "FTE multiplier", "Presence FTE per post"])
    cov_start = r
    for cov, m in R.COVERAGE_MULTIPLIER.items():
        ws.cell(r, 1, cov).font = BOLD
        ws.cell(r, 2, m).font = BASE
        ws.cell(r, 3, R.PRESENCE_FTE_PER_POST[cov]).font = BASE
        r += 1
    names["CoverageTable"] = f"Rates!$A${cov_start}:$C${r-1}"
    r += 1

    # --- Spares ---
    section("SPARES STRATEGY", ["Strategy", "% of hardware value"])
    sp_start = r
    for s, pct in R.SPARES_PCT.items():
        ws.cell(r, 1, s).font = BOLD
        c = ws.cell(r, 2, pct)
        c.number_format = PCT
        c.font = BASE
        r += 1
    names["SparesTable"] = f"Rates!$A${sp_start}:$B${r-1}"
    r += 1

    # --- Survey ---
    section("SURVEY TYPE", ["Type", "Days per AP"])
    sv_start = r
    for t, d in R.SURVEY_DAYS_PER_AP.items():
        ws.cell(r, 1, t).font = BOLD
        ws.cell(r, 2, d).font = BASE
        r += 1
    names["SurveyTable"] = f"Rates!$A${sv_start}:$B${r-1}"
    r += 1

    # --- Scalars ---
    section("MODEL CONSTANTS", ["Constant", "Value"])
    scalars = [
        ("AP_INSTALL_DAYS", R.AP_INSTALL_DAYS),
        ("SWITCH_INSTALL_DAYS", R.SWITCH_INSTALL_DAYS),
        ("MOBILISATION_DAYS_PER_SITE", R.MOBILISATION_DAYS_PER_SITE),
        ("DESIGN_DAYS_FIXED", R.DESIGN_DAYS_FIXED),
        ("DESIGN_DAYS_PER_SITE", R.DESIGN_DAYS_PER_SITE),
        ("PM_OVERHEAD_PCT", R.PM_OVERHEAD_PCT),
        ("OOH_UPLIFT", R.OOH_UPLIFT),
        ("AP_INCIDENT_RATE_PA", R.AP_INCIDENT_RATE_PA),
        ("SWITCH_INCIDENT_RATE_PA", R.SWITCH_INCIDENT_RATE_PA),
        ("HOURS_PER_INCIDENT", R.HOURS_PER_INCIDENT),
        ("PRODUCTIVE_HOURS_PA", R.PRODUCTIVE_HOURS_PA),
        ("WORKING_DAYS_PA", R.WORKING_DAYS_PA),
        ("APS_PER_NOC_FTE", R.APS_PER_NOC_FTE),
        ("SD_CONTACTS_PER_USER_PA", R.SD_CONTACTS_PER_USER_PA),
        ("SD_MINUTES_PER_CONTACT", R.SD_MINUTES_PER_CONTACT),
        ("PORTS_PER_SWITCH", R.PORTS_PER_SWITCH),
        ("USERS_PER_PORT", R.USERS_PER_PORT),
        ("AP_UNIT_COST", R.AP_UNIT_COST),
        ("SWITCH_UNIT_COST", R.SWITCH_UNIT_COST),
        ("TOOLING_PER_SITE", R.TOOLING_PER_SITE),
        ("OVERHEAD_PCT", R.OVERHEAD_PCT),
        ("SITES_PER_POST", R.SITES_PER_POST),
    ]
    for nm, val in scalars:
        ws.cell(r, 1, nm).font = BASE
        ws.cell(r, 2, val).font = BOLD
        names[nm] = _abs("Rates", r, 2)
        r += 1

    return ws


# ===========================================================================
# Input sheet
# ===========================================================================

def build_input(wb: Workbook):
    ws = wb.create_sheet("Input", 0)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 46
    ws.sheet_view.showGridLines = False

    ws["A1"] = "GP-20  ·  SOLUTION & COST ARCHITECT"
    ws["A1"].font = TITLE_F
    ws["A1"].fill = HDR_FILL
    ws.merge_cells("A1:C1")
    ws["A2"] = "Managed LAN / WLAN rollout with field-support wrap"
    ws["A2"].font = Font(bold=True, size=10, color=ACCENT)
    ws["A3"] = "ILLUSTRATIVE MODEL — synthetic rates, not NSC pricing. Yellow cells are inputs."
    ws["A3"].font = MUTED

    r = 5

    def header(t):
        nonlocal r
        for col in (1, 2, 3):
            ws.cell(r, col).fill = SUB_FILL
        c = ws.cell(r, 1, t)
        c.font = SUB_F
        r += 1

    def field(label, key, value, fmt=None, note="", validation=None):
        nonlocal r
        ws.cell(r, 1, label).font = BASE
        c = ws.cell(r, 2, value)
        c.fill = INPUT_FILL
        c.border = BOX
        c.font = BOLD
        c.alignment = Alignment(horizontal="center")
        if fmt:
            c.number_format = fmt
        if note:
            ws.cell(r, 3, note).font = MUTED
        if validation:
            dv = DataValidation(type="list", formula1=f'"{",".join(validation)}"', allow_blank=False)
            ws.add_data_validation(dv)
            dv.add(c)
        names[key] = _abs("Input", r, 2)
        r += 1

    def computed(label, key, formula, fmt=None, note=""):
        nonlocal r
        ws.cell(r, 1, label).font = BOLD
        c = ws.cell(r, 2, formula)
        c.font = BOLD
        c.alignment = Alignment(horizontal="center")
        if fmt:
            c.number_format = fmt
        if note:
            ws.cell(r, 3, note).font = MUTED
        names[key] = _abs("Input", r, 2)
        r += 1

    header("ENGAGEMENT")
    field("Client name", "in_client", "Aurelian Global (illustrative)")
    field("RFP reference", "in_rfpref", "RFP-2026-114")
    field("Contract term (years)", "in_term", 5)
    field("Rollout duration (months)", "in_rollout", 9)
    r += 1

    header("ESTATE")
    field("Sites — Small (≤50 users)", "in_s_small", 62)
    field("Sites — Medium (51–250)", "in_s_medium", 28)
    field("Sites — Large (251–1000)", "in_s_large", 9)
    field("Sites — Campus (>1000)", "in_s_campus", 2)
    computed("Total sites", "SitesTotal",
             f"=SUM({names['in_s_small'].split('!')[1]}:{names['in_s_campus'].split('!')[1]})".replace("$", ""),
             "#,##0")
    field("Total end users", "in_users", "", "#,##0", "Leave blank to derive from site bands")
    field("Access switch count", "in_switches", "", "#,##0", "Leave blank to derive from user count")
    r += 1

    header("COUNTRY MIX  (shares of estate)")
    first_mix = r
    mix_defaults = {"UK": 0.45, "DE": 0.25, "FR": 0.15, "NL": 0.10, "PL": 0.05}
    for c in R.SUPPORTED_COUNTRIES:
        field(c, f"in_mix_{c}", mix_defaults.get(c, 0.0), PCT)
    last_mix = r - 1
    names["MixRange"] = f"Input!$B${first_mix}:$B${last_mix}"
    computed("Total (auto-normalised if ≠ 100%)", "MixTotal",
             f"=SUM(B{first_mix}:B{last_mix})", PCT)
    r += 1

    header("DEPLOYMENT")
    field("Survey type", "in_survey", "hybrid", None, "predictive / hybrid / onsite",
          list(R.SURVEY_DAYS_PER_AP))
    field("Install window", "in_window", "business_hours", None,
          "out_of_hours applies a ×1.35 uplift", ["business_hours", "out_of_hours"])
    r += 1

    header("SERVICE  (the field-support wrap)")
    field("SLA tier", "in_sla", "gold", None, "gold / platinum commit to on-site response",
          list(R.SLA_MULTIPLIER))
    field("Coverage window", "in_coverage", "24x7", None, "Drives shift-rota headcount",
          list(R.COVERAGE_MULTIPLIER))
    field("Spares strategy", "in_spares", "regional", None, "", list(R.SPARES_PCT))
    field("Monitoring included", "in_monitoring", "TRUE", None, "", ["TRUE", "FALSE"])
    field("Service desk included", "in_servicedesk", "TRUE", None, "", ["TRUE", "FALSE"])
    r += 1

    header("COMMERCIALS")
    field("Target margin %", "in_margin", R.DEFAULT_MARGIN_PCT, PCT)
    field("Contingency %", "in_contingency", R.DEFAULT_CONTINGENCY_PCT, PCT)
    field("Annual indexation %", "in_indexation", R.DEFAULT_INDEXATION_PCT, PCT)

    names["SitesRange"] = (f"Input!{names['in_s_small'].split('!')[1]}:"
                           f"{names['in_s_campus'].split('!')[1].replace('$B$', '$B$')}")
    return ws


# ===========================================================================
# Model sheet
# ===========================================================================

def build_model(wb: Workbook):
    ws = wb.create_sheet("Model")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    for c in "CDEFGH":
        ws.column_dimensions[c].width = 15
    ws.sheet_view.showGridLines = False

    ws["A1"] = "MODEL — live calculation chain"
    ws["A1"].font = TITLE_F
    ws["A1"].fill = HDR_FILL
    ws.merge_cells("A1:H1")

    N = names
    r = 3

    def header(t):
        nonlocal r
        for col in range(1, 9):
            ws.cell(r, col).fill = SUB_FILL
        ws.cell(r, 1, t).font = SUB_F
        r += 1

    def calc(label, key, formula, fmt=None, note=""):
        nonlocal r
        ws.cell(r, 1, label).font = BASE
        c = ws.cell(r, 2, formula)
        c.font = BOLD
        if fmt:
            c.number_format = fmt
        if note:
            ws.cell(r, 3, note).font = MUTED
        N[key] = _abs("Model", r, 2)
        r += 1

    header("ESTATE DERIVATION")
    calc("Total APs", "ApsTotal",
         f"=SUMPRODUCT({N['SitesRange']},{N['ApsPerSite']})", "#,##0")
    calc("Total end users", "Users",
         f"=IF({N['in_users']}=\"\",SUMPRODUCT({N['SitesRange']},{N['UsersPerSite']}),{N['in_users']})",
         "#,##0", "RFP value if given, else derived from bands")
    calc("Access switches", "Switches",
         f"=IF({N['in_switches']}=\"\",MAX({N['SitesTotal']},"
         f"CEILING({N['Users']}*{N['USERS_PER_PORT']}/{N['PORTS_PER_SWITCH']},1)),{N['in_switches']})",
         "#,##0")
    calc("Hardware value", "HardwareValue",
         f"={N['ApsTotal']}*{N['AP_UNIT_COST']}+{N['Switches']}*{N['SWITCH_UNIT_COST']}", GBP)
    r += 1

    header("RATE LOOKUPS")
    calc("Survey days per AP", "SurveyRate",
         f"=VLOOKUP({N['in_survey']},{N['SurveyTable']},2,FALSE)", "0.000")
    calc("SLA FTE multiplier", "SlaMult",
         f"=VLOOKUP({N['in_sla']},{N['SlaTable']},2,FALSE)", NUM2)
    calc("SLA on-site response", "SlaResponse",
         f"=VLOOKUP({N['in_sla']},{N['SlaTable']},3,FALSE)")
    calc("SLA triggers coverage floor", "SlaFloor",
         f"=VLOOKUP({N['in_sla']},{N['SlaTable']},4,FALSE)", "0")
    calc("Coverage FTE multiplier", "CovMult",
         f"=VLOOKUP({N['in_coverage']},{N['CoverageTable']},2,FALSE)", NUM2)
    calc("Presence FTE per post", "PresencePerPost",
         f"=VLOOKUP({N['in_coverage']},{N['CoverageTable']},3,FALSE)", NUM2)
    calc("Spares % of hardware", "SparesPct",
         f"=VLOOKUP({N['in_spares']},{N['SparesTable']},2,FALSE)", PCT)
    calc("Blended field day rate", "BlendedField",
         f"=SUMPRODUCT({N['MixRange']},{N['FieldRates']})/{N['MixTotal']}", GBP)
    calc("Blended SDM day rate", "BlendedSdm",
         f"=SUMPRODUCT({N['MixRange']},{N['SdmRates']})/{N['MixTotal']}", GBP)
    calc("Blended travel per incident", "BlendedTravel",
         f"=SUMPRODUCT({N['MixRange']},{N['TravelRates']})/{N['MixTotal']}", GBP)
    r += 1

    header("DEPLOYMENT  (one-off)")
    calc("Survey days", "DepSurvey", f"={N['ApsTotal']}*{N['SurveyRate']}", NUM1)
    calc("Install days", "DepInstall",
         f"=({N['ApsTotal']}*{N['AP_INSTALL_DAYS']}+{N['Switches']}*{N['SWITCH_INSTALL_DAYS']}"
         f"+{N['SitesTotal']}*{N['MOBILISATION_DAYS_PER_SITE']})"
         f"*IF({N['in_window']}=\"out_of_hours\",{N['OOH_UPLIFT']},1)", NUM1)
    calc("Design days", "DepDesign",
         f"={N['DESIGN_DAYS_FIXED']}+{N['SitesTotal']}*{N['DESIGN_DAYS_PER_SITE']}", NUM1)
    calc("Project management days", "DepPm",
         f"=({N['DepSurvey']}+{N['DepInstall']}+{N['DepDesign']})*{N['PM_OVERHEAD_PCT']}", NUM1)
    calc("Total deployment days", "DepDays",
         f"={N['DepSurvey']}+{N['DepInstall']}+{N['DepDesign']}+{N['DepPm']}", NUM1)
    calc("Deployment labour", "DepLabour",
         f"=({N['DepSurvey']}+{N['DepInstall']})*{N['BlendedField']}"
         f"+({N['DepDesign']}+{N['DepPm']})*{N['BlendedSdm']}", GBP)
    calc("Deployment hardware", "DepHardware", f"={N['HardwareValue']}", GBP)
    calc("Deployment tooling", "DepTooling",
         f"={N['SitesTotal']}*{N['TOOLING_PER_SITE']}", GBP)
    calc("Deployment overhead", "DepOverhead",
         f"={N['DepLabour']}*{N['OVERHEAD_PCT']}", GBP)
    calc("Deployment contingency", "DepContingency",
         f"=({N['DepLabour']}+{N['DepHardware']}+{N['DepTooling']}+{N['DepOverhead']})"
         f"*{N['in_contingency']}", GBP)
    calc("Deployment cost", "DepCost",
         f"={N['DepLabour']}+{N['DepHardware']}+{N['DepTooling']}+{N['DepOverhead']}"
         f"+{N['DepContingency']}", GBP)
    calc("Deployment PRICE", "DepPrice",
         f"={N['DepCost']}/(1-{N['in_margin']})", GBP)
    r += 1

    header("RUN — DEMAND")
    calc("Incidents per annum", "IncidentsPa",
         f"={N['ApsTotal']}*{N['AP_INCIDENT_RATE_PA']}+{N['Switches']}*{N['SWITCH_INCIDENT_RATE_PA']}",
         NUM1)
    calc("Demand-driven field FTE (total)", "DemandFieldFte",
         f"={N['IncidentsPa']}*{N['HOURS_PER_INCIDENT']}/{N['PRODUCTIVE_HOURS_PA']}"
         f"*{N['SlaMult']}*{N['CovMult']}", NUM2)
    r += 1

    # --- Per-country coverage floor table ---
    header("RUN — FIELD PRESENCE BY COUNTRY  (max of demand vs coverage floor)")
    hdr_row = r
    for i, h in enumerate(["Country", "Share", "Demand FTE", "Posts", "Floor FTE",
                           "Applied FTE", "Driver", "Labour"], start=1):
        c = ws.cell(r, i, h)
        c.font = HDR_F
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(horizontal="center")
    r += 1
    ctry_start = r
    mix_first = _first_row(N["MixRange"])
    for idx, ctry in enumerate(R.SUPPORTED_COUNTRIES):
        mix_cell = f"Input!$B${mix_first + idx}"
        share = f"({mix_cell}/{N['MixTotal']})"
        ws.cell(r, 1, ctry).font = BOLD
        ws.cell(r, 2, f"={share}").number_format = PCT
        ws.cell(r, 3, f"={N['DemandFieldFte']}*$B{r}").number_format = NUM2
        # A country with no estate needs no standing presence: guard the share
        # before applying MAX(1,...), or an empty country still books one post.
        ws.cell(r, 4, f"=IF(OR({N['SlaFloor']}<>1,$B{r}=0),0,"
                      f"MAX(1,CEILING({N['SitesTotal']}*$B{r}"
                      f"/{N['SITES_PER_POST']},1)))").number_format = "0"
        ws.cell(r, 5, f"=$D{r}*{N['PresencePerPost']}").number_format = NUM2
        ws.cell(r, 6, f"=MAX($C{r},$E{r})").number_format = NUM2
        ws.cell(r, 7, f'=IF($E{r}>$C{r},"coverage floor","incident volume")').font = BASE
        ws.cell(r, 8, f"=$F{r}*INDEX({N['FieldRates']},{idx+1})*{N['WORKING_DAYS_PA']}"
                ).number_format = GBP
        for col in range(1, 9):
            ws.cell(r, col).border = BOX
        r += 1
    ctry_end = r - 1
    names["AppliedFteRange"] = f"Model!$F${ctry_start}:$F${ctry_end}"
    names["CountryLabourRange"] = f"Model!$H${ctry_start}:$H${ctry_end}"
    names["DemandFteRange"] = f"Model!$C${ctry_start}:$C${ctry_end}"
    names["CountryNameRange"] = f"Model!$A${ctry_start}:$A${ctry_end}"
    names["DriverRange"] = f"Model!$G${ctry_start}:$G${ctry_end}"
    r += 1

    header("RUN — RESOURCE & COST")
    calc("Field FTE (applied)", "FieldFte", f"=SUM({N['AppliedFteRange']})", NUM2)
    calc("Surplus FTE above workload", "SurplusFte",
         f"=SUMIF({N['DriverRange']},\"coverage floor\",{N['AppliedFteRange']})"
         f"-SUMIF({N['DriverRange']},\"coverage floor\",{N['DemandFteRange']})", NUM2,
         "Standing presence that exceeds incident demand")
    # UPPER() coerces both a text "TRUE" from the dropdown and a real boolean.
    calc("NOC FTE", "NocFte",
         f'=IF(UPPER({N["in_monitoring"]}&"")="TRUE",'
         f"{N['ApsTotal']}/{N['APS_PER_NOC_FTE']}*{N['CovMult']},0)", NUM2)
    calc("Service desk FTE", "SdFte",
         f'=IF(UPPER({N["in_servicedesk"]}&"")="TRUE",'
         f"{N['Users']}*{N['SD_CONTACTS_PER_USER_PA']}"
         f"*{N['SD_MINUTES_PER_CONTACT']}/60/{N['PRODUCTIVE_HOURS_PA']},0)", NUM2)
    calc("Primary hub share", "HubPrimShare",
         f'=IF({N["in_coverage"]}="24x7",0.65,1)', NUM2, f"Hub: {R.HUB_PRIMARY}")
    calc("Secondary hub share", "HubSecShare",
         f'=IF({N["in_coverage"]}="24x7",0.35,0)', NUM2, f"Hub: {R.HUB_SECONDARY}")
    calc("Field labour", "FieldLabour", f"=SUM({N['CountryLabourRange']})", GBP)
    calc("Central function labour", "CentralLabour",
         f"=(({N['NocFte']}*{N['HubPrimShare']}*VLOOKUP(\"{R.HUB_PRIMARY}\",{N['CountryRates']},3,FALSE))"
         f"+({N['SdFte']}*{N['HubPrimShare']}*VLOOKUP(\"{R.HUB_PRIMARY}\",{N['CountryRates']},4,FALSE))"
         f"+({N['NocFte']}*{N['HubSecShare']}*VLOOKUP(\"{R.HUB_SECONDARY}\",{N['CountryRates']},3,FALSE))"
         f"+({N['SdFte']}*{N['HubSecShare']}*VLOOKUP(\"{R.HUB_SECONDARY}\",{N['CountryRates']},4,FALSE))"
         f")*{N['WORKING_DAYS_PA']}", GBP)
    calc("Delivery labour (pre-SDM)", "LabourPreSdm",
         f"={N['FieldLabour']}+{N['CentralLabour']}", GBP)
    calc("SDM FTE", "SdmFte",
         f"=IF({N['LabourPreSdm']}<1000000,0.5,IF({N['LabourPreSdm']}<3000000,1,"
         f"IF({N['LabourPreSdm']}<7000000,1.8,2.5)))", NUM2, "Banded on delivery labour")
    calc("Run labour (total)", "RunLabour",
         f"={N['LabourPreSdm']}+{N['SdmFte']}*{N['BlendedSdm']}*{N['WORKING_DAYS_PA']}", GBP)
    calc("Travel", "RunTravel", f"={N['IncidentsPa']}*{N['BlendedTravel']}", GBP)
    calc("Spares", "RunSpares", f"={N['HardwareValue']}*{N['SparesPct']}", GBP)
    calc("Tooling (annual refresh)", "RunTooling",
         f"={N['SitesTotal']}*{N['TOOLING_PER_SITE']}*0.25", GBP)
    calc("Overhead", "RunOverhead",
         f"=({N['RunLabour']}+{N['RunTravel']})*{N['OVERHEAD_PCT']}", GBP)
    calc("Contingency", "RunContingency",
         f"=({N['RunLabour']}+{N['RunTravel']}+{N['RunSpares']}+{N['RunTooling']}"
         f"+{N['RunOverhead']})*{N['in_contingency']}", GBP)
    calc("Run cost p.a.", "RunCost",
         f"={N['RunLabour']}+{N['RunTravel']}+{N['RunSpares']}+{N['RunTooling']}"
         f"+{N['RunOverhead']}+{N['RunContingency']}", GBP)
    calc("Run PRICE p.a. (year 1)", "RunPricePa",
         f"={N['RunCost']}/(1-{N['in_margin']})", GBP)
    calc("Run price per month", "RunPricePm", f"={N['RunPricePa']}/12", GBP)
    calc("Total run FTE", "TotalFte",
         f"={N['FieldFte']}+{N['NocFte']}+{N['SdFte']}+{N['SdmFte']}", NUM2)
    r += 1

    header("CONTRACT VALUE")
    yr_hdr = r
    for i, h in enumerate(["Year", "In term?", "Indexed price"], start=1):
        c = ws.cell(r, i, h)
        c.font = HDR_F
        c.fill = PatternFill("solid", fgColor=ACCENT)
    r += 1
    yr_start = r
    for y in range(1, 11):
        ws.cell(r, 1, y).font = BOLD
        ws.cell(r, 2, f'=IF({y}<={N["in_term"]},"yes","-")').font = BASE
        ws.cell(r, 3, f'=IF({y}<={N["in_term"]},{N["RunPricePa"]}'
                      f'*(1+{N["in_indexation"]})^{y - 1},0)').number_format = GBP
        r += 1
    yr_end = r - 1
    names["YearRange"] = f"Model!$C${yr_start}:$C${yr_end}"
    r += 1
    calc("Total run over term", "RunTotal", f"=SUM({N['YearRange']})", GBP)
    calc("TCV", "Tcv", f"={N['DepPrice']}+{N['RunTotal']}", GBP)

    return ws


# ===========================================================================
# Output sheet
# ===========================================================================

def build_output(wb: Workbook):
    ws = wb.create_sheet("Output")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40
    ws.sheet_view.showGridLines = False
    N = names

    ws["A1"] = "SOLUTION & COST SUMMARY"
    ws["A1"].font = TITLE_F
    ws["A1"].fill = HDR_FILL
    ws.merge_cells("A1:D1")
    ws["A2"] = f"={N['in_client']}"
    ws["A2"].font = Font(bold=True, size=12, color=NAVY)
    ws["B2"] = f"={N['in_rfpref']}"
    ws["B2"].font = MUTED
    ws["A3"] = "ILLUSTRATIVE MODEL — synthetic rates, not NSC pricing."
    ws["A3"].font = MUTED
    ws["A3"].fill = WARN_FILL
    ws.merge_cells("A3:D3")

    r = 5

    def sect(t):
        nonlocal r
        for col in range(1, 5):
            ws.cell(r, col).fill = SUB_FILL
        ws.cell(r, 1, t).font = SUB_F
        r += 1

    def line(label, formula, fmt=None, big=False, note=""):
        nonlocal r
        ws.cell(r, 1, label).font = BOLD if big else BASE
        c = ws.cell(r, 2, formula)
        c.font = Font(bold=True, size=12, color=NAVY) if big else BOLD
        if fmt:
            c.number_format = fmt
        if big:
            c.fill = OUT_FILL
        if note:
            ws.cell(r, 4, note).font = MUTED
        r += 1

    sect("SOLUTION SHAPE")
    line("Total sites", f"={N['SitesTotal']}", "#,##0")
    line("Access points", f"={N['ApsTotal']}", "#,##0")
    line("Access switches", f"={N['Switches']}", "#,##0")
    line("End users", f"={N['Users']}", "#,##0")
    line("SLA tier / response", f"={N['in_sla']}&\" — \"&{N['SlaResponse']}")
    line("Coverage window", f"={N['in_coverage']}")
    line("Spares strategy", f"={N['in_spares']}")
    r += 1

    sect("DEPLOYMENT  (one-off)")
    line("Survey days", f"={N['DepSurvey']}", NUM1)
    line("Install days", f"={N['DepInstall']}", NUM1)
    line("Design days", f"={N['DepDesign']}", NUM1)
    line("Project management days", f"={N['DepPm']}", NUM1)
    line("Total effort (days)", f"={N['DepDays']}", NUM1, big=True)
    line("Hardware", f"={N['DepHardware']}", GBP)
    line("One-off price", f"={N['DepPrice']}", GBP, big=True,
         note=f"Delivered over {'{'}rollout{'}'} months")
    ws.cell(r - 1, 4, "").font = MUTED
    r += 1

    sect("RUN — RESOURCE BY LOCATION")
    for i, h in enumerate(["Country", "Role", "FTE", "Sizing driver"], start=1):
        c = ws.cell(r, i, h)
        c.font = HDR_F
        c.fill = PatternFill("solid", fgColor=ACCENT)
    r += 1
    mix_first = _first_row(N["MixRange"])
    ctry_first = _first_row(N["AppliedFteRange"])
    for idx, ctry in enumerate(R.SUPPORTED_COUNTRIES):
        mr = ctry_first + idx
        ws.cell(r, 1, ctry).font = BOLD
        ws.cell(r, 2, "Field Engineer").font = BASE
        ws.cell(r, 3, f"=Model!$F${mr}").number_format = NUM2
        ws.cell(r, 4, f"=Model!$G${mr}").font = BASE
        for col in range(1, 5):
            ws.cell(r, col).border = BOX
        r += 1
    for role, key, hub in [("NOC Engineer", "NocFte", R.HUB_PRIMARY),
                           ("Service Desk", "SdFte", R.HUB_PRIMARY),
                           ("Service Delivery Manager", "SdmFte", None)]:
        ws.cell(r, 1, hub or "Lead country").font = BOLD
        ws.cell(r, 2, role).font = BASE
        ws.cell(r, 3, f"={N[key]}").number_format = NUM2
        ws.cell(r, 4, "central" if hub else "contract band").font = BASE
        for col in range(1, 5):
            ws.cell(r, col).border = BOX
        r += 1
    line("Total run FTE", f"={N['TotalFte']}", NUM2, big=True)
    r += 1

    sect("RUN — ANNUAL COST BUILD")
    for lbl, key in [("Labour", "RunLabour"), ("Travel", "RunTravel"),
                     ("Spares", "RunSpares"), ("Tooling", "RunTooling"),
                     ("Overhead", "RunOverhead"), ("Contingency", "RunContingency")]:
        line(lbl, f"={N[key]}", GBP)
    line("Annual price (year 1)", f"={N['RunPricePa']}", GBP, big=True)
    line("Monthly price (year 1)", f"={N['RunPricePm']}", GBP)
    r += 1

    sect("CONTRACT VALUE")
    line("Term (years)", f"={N['in_term']}", "0")
    line("One-off", f"={N['DepPrice']}", GBP)
    line("Run over term (indexed)", f"={N['RunTotal']}", GBP)
    line("TOTAL CONTRACT VALUE", f"={N['Tcv']}", GBP, big=True)
    r += 1

    sect("ARCHITECT'S NOTE")
    note = ws.cell(r, 1, f'=IF({N["SurplusFte"]}>0.5,'
                         f'"Coverage floor is driving cost: "&TEXT({N["SurplusFte"]},"0.0")&'
                         f'" FTE of standing presence exceeds incident workload. '
                         f'Consider regional pooling or a relaxed tier in low-density countries.",'
                         f'"All locations are sized by incident volume.")')
    note.font = Font(size=10, color=NAVY, italic=True)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)
    ws.row_dimensions[r].height = 30

    return ws


# ===========================================================================

def main(out_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    build_input(wb)      # Input first so names exist
    build_rates(wb)
    build_model(wb)
    build_output(wb)

    # Register defined names for the scalar constants (readable formulas)
    for nm, addr in names.items():
        if nm.isupper():
            try:
                wb.defined_names.add(DefinedName(nm, attr_text=addr))
            except Exception:
                pass

    wb.move_sheet("Output", offset=-3)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Written: {out_path}")
    print(f"Sheets:  {wb.sheetnames}")


if __name__ == "__main__":
    default = (Path(__file__).resolve().parent.parent.parent
               / "assets" / "WLAN_Cost_Estimator.xlsx")
    out = sys.argv[1] if len(sys.argv) > 1 else str(default)
    main(out)
