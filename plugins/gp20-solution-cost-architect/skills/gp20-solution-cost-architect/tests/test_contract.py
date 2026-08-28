"""
Conformance suite — the gate every model pack must pass.

This is the thing that keeps eight packs honest. It knows nothing about any
particular offering: it discovers whatever packs are installed, runs each
through the same assertions, and drives the generic artefact writers with the
output. A pack that passes here can be consumed by the whole skill; one that
fails cannot ship.

Each pack supplies its own exercise cases in `models/<name>/cases.py` as
CASES = [(label, params), ...]. Every case must be valid input.

Run:  python tests/test_contract.py
"""

from __future__ import annotations

# --- Windows console safety -------------------------------------------------
import sys as _sys
for _s in (_sys.stdout, _sys.stderr):
    try:
        if _s and getattr(_s, "encoding", "").lower().replace("-", "") != "utf8":
            _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
# ---------------------------------------------------------------------------

import importlib
import json
import subprocess
import sys
import tempfile
from pathlib import Path

SKILL_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SKILL_ROOT))

from core import registry
from core.contract import ValidationError, verify

passed = failed = 0


def check(label: str, ok: bool, detail: str = ""):
    global passed, failed
    if ok:
        passed += 1
        print(f"    PASS  {label}")
    else:
        failed += 1
        print(f"    FAIL  {label}   {detail}")


def load_cases(folder: str) -> list[tuple[str, dict]]:
    try:
        mod = importlib.import_module(f"models.{folder}.cases")
    except ModuleNotFoundError:
        return []
    return list(getattr(mod, "CASES", []))


def exercise_pack(folder: str, tmp: Path) -> None:
    pack = importlib.import_module(f"models.{folder}.pack")
    man = pack.MANIFEST
    print(f"\n  {man.name}  [{man.key}]")

    # --- manifest hygiene ---------------------------------------------------
    check("manifest key is slug-like",
          man.key and man.key == man.key.lower() and " " not in man.key)
    check("manifest declares detection signals", len(man.detect) >= 3,
          f"{len(man.detect)} signal(s)")
    check("manifest declares price-material parameters", bool(man.material))
    check("every gap option states a commercial consequence",
          all(o.consequence for g in man.gaps for o in g.options))
    check("every gap declares at least two options",
          all(len(g.options) >= 2 for g in man.gaps),
          "a question with one answer is not a question")

    gap_params = {g.param for g in man.gaps}
    check("gaps and material parameters overlap",
          bool(gap_params & set(man.material)) or not man.gaps,
          "asking about things that do not move the price wastes the user's time")

    # --- rejects bad input --------------------------------------------------
    try:
        pack.estimate({})
        check("empty input rejected", False, "no error raised")
    except ValidationError:
        check("empty input rejected", True)
    except Exception as exc:                                    # noqa: BLE001
        check("empty input rejected", False,
              f"wrong exception {type(exc).__name__}")

    # --- cases --------------------------------------------------------------
    cases = load_cases(folder)
    check("pack supplies exercise cases", bool(cases),
          "add models/<pack>/cases.py with CASES = [(label, params), ...]")

    for label, params in cases:
        result = pack.estimate(params)
        problems = verify(result, man)
        check(f"case '{label}' conforms to the contract", not problems,
              "; ".join(problems[:3]))

        s = result["summary"]
        check(f"case '{label}' produces a positive TCV", s["tcv"] > 0)
        check(f"case '{label}' indexes later years",
              len(s["yearly"]) < 2 or s["yearly"][-1]["price"] >= s["yearly"][0]["price"])

        # provenance must survive
        tagged = {a["parameter"] for a in result["assumptions"]}
        check(f"case '{label}' records provenance for supplied params",
              all(k in tagged or k.startswith("_") or isinstance(params[k], dict)
                  for k in params),
              f"untracked: {sorted(set(params) - tagged - {'_sources', '_pack'})[:4]}")

        # cost must sit below price, and the two must bracket the margin
        check(f"case '{label}' keeps cost below price",
              s["total_cost"] < s["tcv"] and s["one_off_cost"] <= s["one_off_price"])

        # --- bill of materials ---------------------------------------------
        bom = result.get("bom") or []
        check(f"case '{label}' produces a bill of materials", bool(bom),
              "every pack buys something — even a desk buys licences")
        if bom:
            check(f"case '{label}' BoM lines all state a unit and a quantity",
                  all(b["unit"] and b["qty"] != 0 for b in bom))
            tagged_lines = [b for b in bom if b.get("rolls_into")]
            check(f"case '{label}' BoM reconciles to the cost model",
                  bool(tagged_lines),
                  "at least one line must name the cost line it belongs to")
            # verify() already proves the sums agree; this proves the BoM is
            # material rather than a token line, which verify() cannot judge.
            materials = sum(abs(b["extended_cost"]) for b in bom)
            check(f"case '{label}' BoM is material to the cost",
                  materials > 0.001 * s["total_cost"],
                  f"BoM totals {materials:,.0f} against "
                  f"{s['total_cost']:,.0f} of cost")

    # --- writers can consume it --------------------------------------------
    if cases:
        result = pack.estimate(cases[0][1])
        rj = tmp / f"{folder}_result.json"
        rj.write_text(json.dumps(result, indent=2), encoding="utf-8")

        from core.write_pricing_form import build as build_form
        xlsx = tmp / f"{folder}_form.xlsx"
        build_form(result, str(xlsx))
        check("generic pricing form writer consumes it",
              xlsx.exists() and xlsx.stat().st_size > 5000)

        from openpyxl import load_workbook
        check("pricing form carries a Bill of Materials sheet",
              "Bill of Materials" in load_workbook(xlsx).sheetnames)

        from core.write_deck import build as build_deck
        pptx = tmp / f"{folder}_deck.pptx"
        build_deck(result, str(pptx))
        check("generic deck writer consumes it",
              pptx.exists() and pptx.stat().st_size > 20000)

        from core.write_cost_model import build as build_model
        model = tmp / f"{folder}_cost_model.xlsx"
        build_model(result, str(model))
        check("generic cost model writer consumes it",
              model.exists() and model.stat().st_size > 5000)

        # --- documented CLI pipeline ---------------------------------------
        pj = tmp / f"{folder}_params.json"
        pj.write_text(json.dumps(cases[0][1]), encoding="utf-8")
        out = tmp / f"{folder}_cli.json"
        with open(out, "w", encoding="utf-8") as fh:
            proc = subprocess.run(
                [sys.executable, str(SKILL_ROOT / "core" / "run_estimate.py"),
                 "--pack", man.key, str(pj)],
                stdout=fh, stderr=subprocess.PIPE, text=True, timeout=120)
        check("run_estimate.py exits cleanly", proc.returncode == 0,
              proc.stderr[-160:])
        try:
            piped = json.loads(out.read_text(encoding="utf-8"))
            check("redirected stdout is pure JSON", True)
            check("CLI result matches the API",
                  abs(piped["summary"]["tcv"] - result["summary"]["tcv"]) < 1.0)
        except json.JSONDecodeError as exc:
            check("redirected stdout is pure JSON", False, str(exc))


def main() -> int:
    print("=" * 68)
    print("  Model pack conformance")
    print("=" * 68)

    folders = registry.pack_keys()
    if not folders:
        print("No packs found.")
        return 1
    print(f"Discovered {len(folders)} pack(s): {', '.join(folders)}")

    tmp = Path(tempfile.mkdtemp(prefix="gp20_contract_"))
    for folder in folders:
        exercise_pack(folder, tmp)

    # --- cross-pack invariants ---------------------------------------------
    print("\n  Cross-pack")
    keys = [importlib.import_module(f"models.{f}.pack").MANIFEST.key for f in folders]
    check("pack keys are unique", len(keys) == len(set(keys)), str(keys))

    # Detection signals must be unique *within* a kind, not across kinds. A
    # tender saying "break-fix" may be asking for a field-service lot or for a
    # whole managed offering — that ambiguity is real, and the skill resolves it
    # by asking whether the tender is lotted. Forcing the signals apart would
    # mean deleting a legitimate phrase from one of them.
    for kind in ("tower", "offering"):
        detects: dict[str, list[str]] = {}
        members = [f for f in folders
                   if importlib.import_module(f"models.{f}.pack").MANIFEST.kind == kind]
        for f in members:
            for sig in importlib.import_module(f"models.{f}.pack").MANIFEST.detect:
                detects.setdefault(sig.lower(), []).append(f)
        clashes = {s: p for s, p in detects.items() if len(p) > 1}
        check(f"no detection signal claimed by two {kind} packs", not clashes,
              f"ambiguous: {list(clashes)[:3]}")

    kinds = {importlib.import_module(f"models.{f}.pack").MANIFEST.kind
             for f in folders}
    check("every pack declares a valid kind", kinds <= {"tower", "offering"},
          str(kinds))

    print("\n" + "=" * 68)
    print(f"  {passed} passed, {failed} failed")
    print(f"  Artefacts: {tmp}")
    print("=" * 68)
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
