"""
Parity test — proves a pack's Excel workbook and its Python model agree.

Optional per pack: only offerings that ship a `workbook.py` are checked. A
workbook exists so a Solution Shaper can inspect and change the logic without
the AI in the picture; this test is what stops the two implementations drifting
apart once that is true.

Writes a scenario into the workbook's Input sheet, recalculates it headlessly
with LibreOffice, then compares the recalculated Output against estimator.py.

Run:  python3 test_parity.py
Exit: 0 = all scenarios match, 1 = drift detected
"""



from __future__ import annotations

# --- Windows console safety -------------------------------------------------
# Windows defaults stdout to cp1252, which cannot encode the currency symbols,
# em-dashes and comparison operators this model emits. Force UTF-8 so the
# caller never has to remember `python -X utf8`.
import sys as _sys
for _s in (_sys.stdout, _sys.stderr):
    try:
        if _s and getattr(_s, "encoding", "").lower().replace("-", "") != "utf8":
            _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
# ---------------------------------------------------------------------------


import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

SKILL_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(SKILL_ROOT))

from models.managed_lan import rates as R                      # noqa: E402
from models.managed_lan.pack import estimate                   # noqa: E402
from models.managed_lan.workbook import main as build_wb       # noqa: E402

TOLERANCE = 0.005  # 0.5% — absorbs Excel/Python float and rounding differences

SCENARIOS = [
    ("Gold / 24x7, 5 countries", {
        "client_name": "Aurelian Global (illustrative)", "term_years": 5,
        "sites": {"small": 62, "medium": 28, "large": 9, "campus": 2},
        "country_mix": {"UK": 0.45, "DE": 0.25, "FR": 0.15, "NL": 0.10, "PL": 0.05},
        "survey_type": "hybrid", "sla_tier": "gold", "coverage": "24x7",
        "spares_strategy": "regional",
    }),
    ("Silver / 8x5, UK only", {
        "client_name": "Test B", "term_years": 3,
        "sites": {"small": 20, "medium": 5},
        "country_mix": {"UK": 1.0},
        "survey_type": "predictive", "sla_tier": "silver", "coverage": "8x5",
        "spares_strategy": "central",
    }),
    ("Platinum / 24x7, out-of-hours install", {
        "client_name": "Test C", "term_years": 7,
        "sites": {"medium": 40, "large": 12},
        "country_mix": {"DE": 0.6, "PL": 0.4},
        "survey_type": "onsite", "sla_tier": "platinum", "coverage": "24x7",
        "spares_strategy": "onsite", "install_window": "out_of_hours",
    }),
    ("Bronze / 12x5, no monitoring or service desk", {
        "client_name": "Test D", "term_years": 3,
        "sites": {"small": 150},
        "country_mix": {"PL": 1.0},
        "survey_type": "hybrid", "sla_tier": "bronze", "coverage": "12x5",
        "spares_strategy": "none", "monitoring": False, "service_desk": False,
    }),
]

# Output-sheet label -> path into the estimator result
def _head(r, label):
    return next(i["value"] for i in r["scope"]["headline"] if i["label"] == label)


CHECKS = {
    "Total sites": lambda r: _head(r, "Locations in scope"),
    "Access points": lambda r: _head(r, "Wireless access points"),
    "Access switches": lambda r: _head(r, "Access switches"),
    "End users": lambda r: _head(r, "End users supported"),
    "Total effort (days)": lambda r: r["deployment"]["days_by_role"]["total"],
    "One-off price": lambda r: r["deployment"]["price"],
    "Total run FTE": lambda r: r["run"]["total_fte"],
    "Annual price (year 1)": lambda r: r["run"]["price_pa"],
    "TOTAL CONTRACT VALUE": lambda r: r["summary"]["tcv"],
}


def write_inputs(path: Path, params: dict):
    """Push a scenario into the workbook's Input sheet."""
    wb = load_workbook(path)
    ws = wb["Input"]
    lookup = {}
    for row in ws.iter_rows(min_col=1, max_col=2):
        label = row[0].value
        if isinstance(label, str):
            lookup[label.strip()] = row[1]

    sites = params.get("sites", {})
    lookup["Sites — Small (≤50 users)"].value = sites.get("small", 0)
    lookup["Sites — Medium (51–250)"].value = sites.get("medium", 0)
    lookup["Sites — Large (251–1000)"].value = sites.get("large", 0)
    lookup["Sites — Campus (>1000)"].value = sites.get("campus", 0)

    mix = params.get("country_mix", {"UK": 1.0})
    for c in R.SUPPORTED_COUNTRIES:
        lookup[c].value = mix.get(c, 0.0)

    lookup["Client name"].value = params.get("client_name", "")
    lookup["Contract term (years)"].value = params.get("term_years", R.DEFAULT_TERM_YEARS)
    lookup["Total end users"].value = params.get("user_count", "")
    lookup["Access switch count"].value = params.get("switch_count", "")
    lookup["Survey type"].value = params.get("survey_type", "hybrid")
    lookup["Install window"].value = params.get("install_window", "business_hours")
    lookup["SLA tier"].value = params.get("sla_tier", "silver")
    lookup["Coverage window"].value = params.get("coverage", "8x5")
    lookup["Spares strategy"].value = params.get("spares_strategy", "regional")
    lookup["Monitoring included"].value = str(params.get("monitoring", True)).upper()
    lookup["Service desk included"].value = str(params.get("service_desk", True)).upper()
    lookup["Target margin %"].value = params.get("margin_pct", R.DEFAULT_MARGIN_PCT)
    lookup["Contingency %"].value = params.get("contingency_pct", R.DEFAULT_CONTINGENCY_PCT)
    lookup["Annual indexation %"].value = params.get("indexation_pct", R.DEFAULT_INDEXATION_PCT)
    wb.save(path)


def recalc(path: Path, outdir: Path) -> Path:
    subprocess.run(
        ["soffice", "--headless", "--calc", "--convert-to", "xlsx",
         "--outdir", str(outdir), str(path)],
        check=True, capture_output=True, timeout=300,
    )
    return outdir / path.name


def read_output(path: Path) -> dict:
    ws = load_workbook(path, data_only=True)["Output"]
    vals = {}
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        if isinstance(row[0], str) and row[1] is not None:
            vals.setdefault(row[0].strip(), row[1])
    return vals


def main() -> int:
    tmp = Path(tempfile.mkdtemp(prefix="gp20_parity_"))
    master = tmp / "model.xlsx"
    build_wb(str(master))

    failures = 0
    print(f"{'Scenario':<44} {'Check':<26} {'Excel':>16} {'Python':>16}  Δ")
    print("-" * 112)

    for name, params in SCENARIOS:
        book = tmp / f"{abs(hash(name))}.xlsx"
        shutil.copy(master, book)
        write_inputs(book, params)
        out_dir = tmp / f"out_{abs(hash(name))}"
        out_dir.mkdir(exist_ok=True)
        recalced = recalc(book, out_dir)
        excel = read_output(recalced)
        py = estimate(params)

        first = True
        for label, getter in CHECKS.items():
            x = excel.get(label)
            p = getter(py)
            if x is None:
                print(f"{name if first else '':<44} {label:<26} {'MISSING':>16}")
                failures += 1
                first = False
                continue
            denom = max(abs(p), 1.0)
            delta = abs(float(x) - float(p)) / denom
            ok = delta <= TOLERANCE
            if not ok:
                failures += 1
            mark = "ok" if ok else "*** DRIFT"
            print(f"{name if first else '':<44} {label:<26} "
                  f"{float(x):>16,.2f} {float(p):>16,.2f}  {delta:>6.3%} {mark}")
            first = False
        print()

    shutil.rmtree(tmp, ignore_errors=True)
    if failures:
        print(f"FAILED — {failures} mismatch(es)")
        return 1
    print(f"PASSED — Excel and Python agree across "
          f"{len(SCENARIOS)} scenarios × {len(CHECKS)} checks "
          f"(tolerance {TOLERANCE:.1%})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
