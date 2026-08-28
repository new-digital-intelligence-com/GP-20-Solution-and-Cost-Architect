#!/usr/bin/env python3
"""
Cost-model workbook parity — every pack, every case.

The cost model is a live workbook: costs are values written by the pack, but
margin, price, the term schedule and contract value are Excel formulas. That
makes it useful to finance and dangerous to us — a formula that disagrees with
the Python is a wrong number in a document with our name on it, and nothing in
the Python test suite would ever see it.

So this recalculates each workbook headlessly with LibreOffice and compares the
formula results against the pack. Every pack, every case, no sampling.

Run:  python tests/test_cost_model.py
"""

from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook                                   # noqa: E402

from core.registry import catalogue, load                            # noqa: E402
from core.write_cost_model import build                              # noqa: E402

TOLERANCE = 0.001          # 0.1% — formulas should agree far closer than this


def recalc(path: Path, outdir: Path) -> Path:
    subprocess.run(
        ["soffice", "--headless", "--calc", "--convert-to", "xlsx",
         "--outdir", str(outdir), str(path)],
        check=True, capture_output=True, timeout=300,
    )
    return outdir / path.name


def read_labels(path: Path) -> dict:
    """Label in column A, first numeric value found to its right."""
    ws = load_workbook(path, data_only=True)["Cost Model"]
    vals: dict[str, float] = {}
    for row in ws.iter_rows(min_col=1, max_col=4, values_only=True):
        if not isinstance(row[0], str):
            continue
        for cell in row[1:]:
            if isinstance(cell, (int, float)):
                vals.setdefault(row[0].strip(), float(cell))
                break
    return vals


CHECKS = {
    "Total one-off cost":       lambda r: r["summary"]["one_off_cost"],
    "One-off price":            lambda r: r["summary"]["one_off_price"],
    "Total annual cost (year 1)": lambda r: r["summary"]["annual_cost_year1"],
    "Annual price (year 1)":    lambda r: r["summary"]["annual_price_year1"],
    "Monthly price (year 1)":   lambda r: r["summary"]["monthly_price_year1"],
    "Total delivery cost over term": lambda r: r["summary"]["total_cost"],
    "Total contract value":     lambda r: r["summary"]["tcv"],
    "Margin":                   lambda r: r["summary"]["margin_value"],
}


def main() -> int:
    if not shutil.which("soffice"):
        print("SKIP — LibreOffice not installed; cannot recalculate formulas.")
        print("      Install libreoffice-calc to run this gate.")
        return 0

    tmp = Path(tempfile.mkdtemp(prefix="gp20_costmodel_"))
    failures = checks = 0

    print(f"{'Pack / case':<52} {'Check':<32} {'Excel':>15} {'Python':>15}")
    print("-" * 118)

    for entry in catalogue():
        key = entry["key"]
        mod = load(key)
        cases = __import__(f"{mod.__name__.rsplit('.', 1)[0]}.cases",
                           fromlist=["CASES"]).CASES
        for idx, (name, params) in enumerate(cases):
            result = mod.estimate(params)
            src = tmp / f"{key}_{idx}.xlsx"
            build(result, str(src))
            out_dir = tmp / f"out_{key}_{idx}"
            out_dir.mkdir(exist_ok=True)
            excel = read_labels(recalc(src, out_dir))

            label = f"{key} — {name}"
            first = True
            for check, getter in CHECKS.items():
                checks += 1
                py = float(getter(result))
                x = excel.get(check)
                if x is None:
                    print(f"{label if first else '':<52} {check:<32} "
                          f"{'MISSING':>15}")
                    failures += 1
                    first = False
                    continue
                drift = abs(x - py) / max(abs(py), 1.0)
                ok = drift <= TOLERANCE
                if not ok:
                    failures += 1
                if not ok or first:
                    mark = "" if ok else "   <-- FAIL"
                    print(f"{label if first else '':<52} {check:<32} "
                          f"{x:>15,.2f} {py:>15,.2f}{mark}")
                    first = False

    print("-" * 118)
    if failures:
        print(f"  {checks - failures}/{checks} checks passed, {failures} FAILED")
    else:
        print(f"  {checks} checks passed — every workbook formula agrees "
              f"with its pack")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
